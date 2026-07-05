import openpyxl, math, json
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
def r(x): return int(math.floor(x+0.5))
D=json.load(open("_data.json"))

def t(hm):
    if not hm: return None
    h,m=map(int,hm.split(":")); return datetime(2000,1,2,h,m)
def shift(base,hrs):
    return (base-timedelta(hours=hrs))

wb=openpyxl.load_workbook("Foundation Schedule.xlsx")
if "As Received" in wb.sheetnames: del wb["As Received"]
ws=wb.create_sheet("As Received")

headers=["DATE","Airline code","Flight Designator","SOBT","Hour","Routing","Traffic Type",
         "Public Terminal","GHA","Seats","Pax","D-4 Pax","D-3 Pax","D-2 Pax","D-4","D-3","D-2"]
hf=PatternFill("solid",fgColor="1F3864"); hfont=Font(name="Cambria",bold=True,color="FFFFFF",size=10)
flag=PatternFill("solid",fgColor="FCE4D6")  # light orange for low-confidence
thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
cf=Font(name="Cambria",size=10)
ws.append(headers)
for c in ws[1]:
    c.fill=hf;c.font=hfont;c.alignment=Alignment("center","center",wrap_text=True);c.border=bd

flagged=[]
for i,(al,fl,sobt,rt,gha,seats,conf) in enumerate(D):
    rownum=i+2
    base=t(sobt)
    if seats is None:
        row=["21-06-2026",al,fl,None,None,rt,"INTERNATIONAL","Terminal 3",gha,None,None,None,None,None,None,None,None]
    else:
        pax=r(0.8*seats); d4=r(0.3*pax); d3=r(0.4*pax); d2=r(0.3*pax)
        row=["21-06-2026",al,fl,base,base.hour,rt,"INTERNATIONAL","Terminal 3",gha,seats,pax,d4,d3,d2,
             shift(base,4),shift(base,3),shift(base,2)]
    ws.append(row)
    if not conf: flagged.append(rownum)
    for c in ws[rownum]:
        c.font=cf; c.border=bd; c.alignment=Alignment("center","center")
    ws.cell(rownum,4).number_format="hh:mm"
    for cc in (15,16,17): ws.cell(rownum,cc).number_format="hh:mm"
    if not conf:
        for c in ws[rownum]: c.fill=flag

n=len(D); last=n+1
# totals row
tr=last+1
ws.cell(tr,3,"TOTAL").font=Font(name="Cambria",bold=True)
for col in (10,11,12,13,14):
    L=openpyxl.utils.get_column_letter(col)
    ws.cell(tr,col,f"=SUM({L}2:{L}{last})").font=Font(name="Cambria",bold=True)
    ws.cell(tr,col).number_format="#,##0"

widths={"A":11,"B":11,"C":15,"D":8,"E":7,"F":13,"G":13,"H":12,"I":10,"J":8,"K":8,"L":8,"M":8,"N":8,"O":8,"P":8,"Q":8}
for k,v in widths.items(): ws.column_dimensions[k].width=v
ws.freeze_panes="A2"

# ---- rename Foundation Schedule -> As Cleaned, rebuild as references ----
if "Foundation Schedule" in wb.sheetnames:
    fs=wb["Foundation Schedule"]; fs.title="As Cleaned"
elif "As Cleaned" in wb.sheetnames:
    fs=wb["As Cleaned"]
else:
    fs=wb.create_sheet("As Cleaned")
# wipe As Cleaned
wb.remove(fs)
ac=wb.create_sheet("As Cleaned",0)
ac.append(headers+["Σ split = Pax?","Pax = 80%·Seats?"])
for c in ac[1]:
    c.fill=hf;c.font=hfont;c.alignment=Alignment("center","center",wrap_text=True);c.border=bd
for i in range(n):
    rr=i+2
    for col in range(1,18):
        L=openpyxl.utils.get_column_letter(col)
        ac.cell(rr,col,f"='As Received'!{L}{rr}")
        ac.cell(rr,col).font=cf; ac.cell(rr,col).border=bd; ac.cell(rr,col).alignment=Alignment("center","center")
    ac.cell(rr,4).number_format="hh:mm"
    for cc in (15,16,17): ac.cell(rr,cc).number_format="hh:mm"
    # checks
    ac.cell(rr,18,f"=IF(K{rr}=\"\",\"\",K{rr}-(L{rr}+M{rr}+N{rr}))")
    ac.cell(rr,19,f"=IF(J{rr}=\"\",\"\",K{rr}-ROUND(0.8*J{rr},0))")
    for cc in (18,19):
        ac.cell(rr,cc).font=cf; ac.cell(rr,cc).border=bd; ac.cell(rr,cc).alignment=Alignment("center","center")
tr2=n+2
ac.cell(tr2,3,"TOTAL").font=Font(name="Cambria",bold=True)
for col in (10,11,12,13,14):
    L=openpyxl.utils.get_column_letter(col)
    ac.cell(tr2,col,f"=SUM({L}2:{L}{n+1})").font=Font(name="Cambria",bold=True)
    ac.cell(tr2,col).number_format="#,##0"
for k,v in widths.items(): ac.column_dimensions[k].width=v
ac.column_dimensions["R"].width=13; ac.column_dimensions["S"].width=14
ac.freeze_panes="A2"

# order: As Cleaned, As Received, GHA Assignment, Counter Times
order=["As Cleaned","As Received","GHA Assignment","Counter Times"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save("Foundation Schedule.xlsx")

# code-side totals check
seats=[d[5] for d in D if d[5] is not None]
pax=[r(0.8*s) for s in seats]
print("flights:",n,"| with seats:",len(seats),"| placeholders:",n-len(seats))
print("TOTAL seats:",sum(seats),"| TOTAL pax:",sum(pax))
print("flagged rows (Excel):",flagged)
print("sheets:",wb.sheetnames)
