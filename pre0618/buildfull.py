import openpyxl, math, json
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as col
from datetime import datetime, timedelta
from collections import Counter, defaultdict
def r(x): return int(math.floor(x+0.5))
D=json.load(open("_data.json"))
def t(hm):
    if not hm: return None
    h,m=map(int,hm.split(":")); return datetime(2000,1,2,h,m)
def shift(b,h): return b-timedelta(hours=h)

OUT="/tmp/Foundation Schedule.xlsx"
hf=PatternFill("solid",fgColor="1F3864"); hfont=Font(name="Cambria",bold=True,color="FFFFFF",size=10)
flag=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
cf=Font(name="Cambria",size=10); bfont=Font(name="Cambria",bold=True,size=10)
def H(ws,row=1,extra=0,ncol=17):
    for c in ws[row][:ncol+extra]:
        c.fill=hf;c.font=hfont;c.alignment=Alignment("center","center",wrap_text=True);c.border=bd

wb=openpyxl.Workbook(); wb.remove(wb.active)
headers=["DATE","Airline code","Flight Designator","SOBT","Hour","Routing","Traffic Type",
         "Public Terminal","GHA","Seats","Pax","D-4 Pax","D-3 Pax","D-2 Pax","D-4","D-3","D-2"]
n=len(D)

# ---------- As Received ----------
ar=wb.create_sheet("As Received"); ar.append(headers); H(ar)
flagged=[]
for i,(al,fl,sobt,rt,gha,seats,conf) in enumerate(D):
    rn=i+2; b=t(sobt)
    if seats is None:
        ar.append(["21-06-2026",al,fl,None,None,rt,"INTERNATIONAL","Terminal 3",gha,None,None,None,None,None,None,None,None])
    else:
        pax=r(0.8*seats); d4=r(0.3*pax); d3=r(0.4*pax); d2=r(0.3*pax)
        ar.append(["21-06-2026",al,fl,b,b.hour,rt,"INTERNATIONAL","Terminal 3",gha,seats,pax,d4,d3,d2,shift(b,4),shift(b,3),shift(b,2)])
    if not conf: flagged.append(rn)
    for c in ar[rn]:
        c.font=cf;c.border=bd;c.alignment=Alignment("center","center")
        if not conf: c.fill=flag
    ar.cell(rn,4).number_format="hh:mm"
    for cc in (15,16,17): ar.cell(rn,cc).number_format="hh:mm"
trA=n+2; ar.cell(trA,3,"TOTAL").font=bfont
for c_ in (10,11,12,13,14):
    L=col(c_); ar.cell(trA,c_,f"=SUM({L}2:{L}{n+1})").font=bfont; ar.cell(trA,c_).number_format="#,##0"
W={"A":11,"B":11,"C":15,"D":8,"E":7,"F":13,"G":13,"H":11,"I":10,"J":7,"K":7,"L":7,"M":7,"N":7,"O":7,"P":7,"Q":7}
for k,v in W.items(): ar.column_dimensions[k].width=v
ar.freeze_panes="A2"

# ---------- As Cleaned (references) ----------
ac=wb.create_sheet("As Cleaned"); ac.append(headers+["Σ split = Pax?","Pax = 80%·Seats?"]); H(ac,extra=2)
for i in range(n):
    rr=i+2
    for c_ in range(1,18):
        L=col(c_); cell=ac.cell(rr,c_,f"='As Received'!{L}{rr}")
        cell.font=cf;cell.border=bd;cell.alignment=Alignment("center","center")
    ac.cell(rr,4).number_format="hh:mm"
    for cc in (15,16,17): ac.cell(rr,cc).number_format="hh:mm"
    ac.cell(rr,18,f'=IF(K{rr}="","",K{rr}-(L{rr}+M{rr}+N{rr}))')
    ac.cell(rr,19,f'=IF(J{rr}="","",K{rr}-ROUND(0.8*J{rr},0))')
    for cc in (18,19): ac.cell(rr,cc).font=cf;ac.cell(rr,cc).border=bd;ac.cell(rr,cc).alignment=Alignment("center","center")
trC=n+2; ac.cell(trC,3,"TOTAL").font=bfont
for c_ in (10,11,12,13,14):
    L=col(c_); ac.cell(trC,c_,f"=SUM({L}2:{L}{n+1})").font=bfont; ac.cell(trC,c_).number_format="#,##0"
for k,v in W.items(): ac.column_dimensions[k].width=v
ac.column_dimensions["R"].width=13; ac.column_dimensions["S"].width=14
ac.freeze_panes="A2"

# ---------- GHA Assignment (majority per airline, flag conflicts) ----------
by=defaultdict(list)
for al,fl,sobt,rt,gha,seats,conf in D:
    if al!="VERIFY" and gha: by[al].append(gha)
ga=wb.create_sheet("GHA Assignment"); ga.append(["Airline Code","GHA (majority)","# Flights","Mixed?"]); H(ga,ncol=4)
rr=2
for al in sorted(by):
    ghas=by[al]; cnt=Counter(ghas); top=cnt.most_common(1)[0][0]; mixed="YES "+str(dict(cnt)) if len(cnt)>1 else ""
    ga.append([al,top,len(ghas),mixed])
    for c in ga[rr]:
        c.font=cf;c.border=bd;c.alignment=Alignment("center","center")
        if mixed: c.fill=flag
    rr+=1
for k,v in {"A":13,"B":16,"C":10,"D":34}.items(): ga.column_dimensions[k].width=v
ga.freeze_panes="A2"

# ---------- Counter Times (all flights, SOBT-4 open / SOBT-1 close) ----------
ct=wb.create_sheet("Counter Times"); ct.append(["Flight #","SOBT","Counter Open (SOBT-4)","Counter Close (SOBT-1)"]); H(ct,ncol=4)
rr=2
for al,fl,sobt,rt,gha,seats,conf in D:
    b=t(sobt)
    if b: ct.append([fl,b,shift(b,4),shift(b,1)])
    else: ct.append([fl,None,None,None])
    for c in ct[rr]: c.font=cf;c.border=bd;c.alignment=Alignment("center","center")
    for cc in (2,3,4): ct.cell(rr,cc).number_format="hh:mm"
    rr+=1
for k,v in {"A":13,"B":9,"C":20,"D":20}.items(): ct.column_dimensions[k].width=v
ct.freeze_panes="A2"

# order
order=["As Cleaned","As Received","GHA Assignment","Counter Times"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.save(OUT)

seats=[d[5] for d in D if d[5] is not None]; pax=[r(0.8*s) for s in seats]
print("flights:",n,"with seats:",len(seats),"placeholders:",n-len(seats))
print("TOTAL seats:",sum(seats),"TOTAL pax:",sum(pax))
print("flagged Excel rows:",flagged)
print("airlines:",len(by),"| mixed-GHA airlines:",[a for a in by if len(set(by[a]))>1])
print("saved to",OUT)
